from operator import truediv
import streamlit as st
import numpy as np
import pandas as pd
import os.path
import sqlite3
import openpyxl
import glob
import time
import datetime
from datetime import datetime as dt


# Streamlit Setting-------------------------------------
st.set_page_config(
    page_title = "Mfg. System",
    page_icon = "//192.168.1.212//アイシス//01_製造_組立進歩システム//00_作業進歩表//ISIS_Logo.png",
    layout="wide",
    initial_sidebar_state="expanded"
)
# ------------------------------------------------------


# Path を指定-------------------------------------------------------------------------------
cd = os.path.dirname(__file__)
Press_csv  = pd.read_csv(filepath_or_buffer="//192.168.1.212/アイシス/00_製造_自動発注システム/10_Press_No/Press_No.csv", 
                         encoding="ANSI", 
                         sep=",")
#Work_schedule_csv = pd.read_csv('C://Users//tani//Desktop//desktop_python2//作業工程表.csv')
Work_schedule_csv = pd.read_csv("//192.168.1.212/アイシス/01_製造_組立進歩システム/01_prog/00_mfg_main/作業工程表.csv")
leader_path = "//192.168.1.212//アイシス//生産管理//会議資料//担当者一覧.xlsx"
BUP_DB_Path = "//192.168.1.212/アイシス/00_製造_自動発注システム/20_DataBase/Auto_Order_Sys.db"

# ページタイトル-----------------------------------------------------------------------------
st.title('Work Sheet')

# -----------------------------------------------------------------------------------------
# Streamlitで遅延理由の記入欄を表示
# -----------------------------
def start_Order_His():
    # ----------------------------------------------------
    # Sidebar
    # ----------------------------------------------------
    
    # 表示ボタンを押さないとページ変更できない（保持）
    SB_change_pages = st.sidebar.radio(f"表示したいページを選択してください。",
                                      ("作業時間登録","経過時間表図化")
                                      )
    SB_button = st.sidebar.button("表示")
    if 'pages' not in st.session_state: 
        st.session_state['pages'] = "作業時間登録" #初期値
    if SB_change_pages == "作業時間登録" and SB_button == True :
        st.session_state['pages'] = "作業時間登録"
    if SB_change_pages == "経過時間表図化" and SB_button == True :
        st.session_state['pages'] = "経過時間表図化"
    st.sidebar.success(f"現在のページは{st.session_state['pages']}です。")

    # ----------------------------------------------------
    # ----------------------------------------------------
    
#作業時間登録ページコード===========================================================================================
    if st.session_state['pages'] == "作業時間登録":
        st.write("<b>作業機種記入欄</b>",unsafe_allow_html=True)
        cols = st.columns((1, 1, 1))

        wb3 = openpyxl.load_workbook(leader_path, data_only=True, read_only=True, keep_vba=True)
        sheet_wb3 = wb3['担当者一覧']
        data_leader = ["-"]
        for i in range(30):
            if sheet_wb3.cell(2 + i, 2).value  != None :
                  data_leader.append(sheet_wb3.cell(2 + i, 2).value)
            else:
                break
        workers_name = cols[0].selectbox("■ 作業者名を選択してください。", data_leader)
        date = cols[1].date_input("■ 作業日")

        cols = st.columns((1, 1, 1))

        selected_Press = cols[0].selectbox(
            "■ 機種の絞り込みです。選択してください。",
            ["-",
            "PLENOX_series",
            "U_series",
            "N_series",
            "S・G_series",
            "ES_series",
            "KIT_series",
            "C_series",
            "DM_series",
            "VIVO_series"
            ]
            )
        list_type = list()
        if selected_Press == "PLENOX_series":
            P_list_offset = 1
            P_list_read   = 12
        elif selected_Press == "U_series":
            P_list_offset = 31
            P_list_read   = 24
        elif selected_Press == "N_series":
            P_list_offset = 61
            P_list_read   = 7
        elif selected_Press == "S・G_series":
            P_list_offset = 91
            P_list_read   = 11
        elif selected_Press == "ES_series":
            P_list_offset = 81
            P_list_read   = 2
        elif selected_Press == "KIT_series":
            P_list_offset = 121
            P_list_read   = 3
        elif selected_Press == "C_series":
            P_list_offset = 111
            P_list_read   = 4
        elif selected_Press == "DM_series":
            P_list_offset = 131
            P_list_read   = 3
        elif selected_Press == "VIVO_series":
            P_list_offset = 21
            P_list_read   = 3
        else :
            P_list_offset = 900
            P_list_read   = 5
            
        readmax = P_list_offset + P_list_read
        list_type = Press_csv[P_list_offset:readmax]["1"]
        selected_Press_type = cols[1].selectbox(
            "ー機種を選択してください。",
            list_type   
            )

        press_No = cols[2].number_input("■ 号機を記入してください。", 1 ,10000, 1)
        
        st.write("<b>作業工程記入欄</b>",unsafe_allow_html=True)
        frame_amount = st.columns((1, 8))[0].number_input("フォームの数", 1 ,10, 1)

        Wrok_time = [0,0,0,0,0,0,0,0,0,0]
        selected_Work_Item = [0,0,0,0,0,0,0,0,0,0]
        selected_Work = [0,0,0,0,0,0,0,0,0,0]
        comment = [0,0,0,0,0,0,0,0,0,0]
        key_number = 1
        
        for FA in range(frame_amount):
            st.markdown('----')
            cols = st.columns((1, 1, 1, 1))
            Wrok_time[FA] = cols[0].time_input("■ 経過時間", datetime.time(00, 00), key = key_number)
            selected_Work_Item[FA] = cols[1].selectbox(
                "■ 作業工程の絞り込みです。選択してください。",
                ["-",
                "ベッド・レッグ",
                "コラム",
                "クラウン・フレーム",
                "クランク",
                "フライホイール",
                "モータ",
                "クラッチ",
                "スライド",
                "ダイナミックバランサー",
                "コネクションユニット",
                "ボルスター",
                "ガイドコラム",
                "配管",
                "仮回し"
                ],
                key = key_number)
            Work_list_code = list()
            if selected_Work_Item[FA] == "ベッド・レッグ":
                W_list_offset = 1
                W_list_read   = 9
            elif selected_Work_Item[FA] == "コラム":
                W_list_offset = 21
                W_list_read   = 3
            elif selected_Work_Item[FA] == "クラウン・フレーム":
                W_list_offset = 31
                W_list_read   = 6
            elif selected_Work_Item[FA] == "クランク":
                W_list_offset = 41
                W_list_read   = 13
            elif selected_Work_Item[FA] == "フライホイール":
                W_list_offset = 61
                W_list_read   = 5
            elif selected_Work_Item[FA] == "モータ":
                W_list_offset = 71
                W_list_read   = 5
            elif selected_Work_Item[FA] == "クラッチ":
                W_list_offset = 81
                W_list_read   = 6
            elif selected_Work_Item[FA] == "スライド":
                W_list_offset = 91
                W_list_read   = 13
            elif selected_Work_Item[FA] == "ダイナミックバランサー":
                W_list_offset = 111
                W_list_read   = 9
            elif selected_Work_Item[FA] == "コネクションユニット":
                W_list_offset = 131
                W_list_read   = 12
            elif selected_Work_Item[FA] == "ボルスター":
                W_list_offset = 151
                W_list_read   = 9
            elif selected_Work_Item[FA] == "ガイドコラム":
                W_list_offset = 171
                W_list_read   = 5
            elif selected_Work_Item[FA] == "配管":
                W_list_offset = 181
                W_list_read   = 12
            elif selected_Work_Item[FA] == "仮回し":
                W_list_offset = 201
                W_list_read   = 1
            else :
                W_list_offset = 900
                W_list_read   = 5

            readmax = W_list_offset + W_list_read
            Work_list_code = Work_schedule_csv[W_list_offset:readmax]["1"]
            selected_Work[FA] = cols[2].selectbox(
                "ー作業工程項目を選択してください。",
                Work_list_code,
                key = key_number)
            comment[FA] = cols[3].text_area("■ コメント", key = key_number)

            key_number += 1

        CB = st.checkbox("✓ 全て記入しました")
        cols = st.columns((1, 4, 8))
        button1 = cols[0].button("登録")

        # 変数の中身(空白)チェック~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        entry = False
        if button1 == True:
            if workers_name != "-" and selected_Press != "-":
                for Re_FA in range(frame_amount):
                    if Wrok_time[Re_FA] != datetime.time(00, 00) and selected_Work_Item[Re_FA] != "-" and comment[Re_FA] != "":
                        entry = True
                    else :
                        entry = False
                        break
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        # ----------------------------------------------------
        if button1 == True and entry == True and CB == True:
            with st.spinner('Wait for it...'):
                time.sleep(2)

            db_sys = Work_sheet_Database()
            db_data = db_sys.get(
                    workers_name,
                    date,
                    selected_Press,
                    selected_Press_type,
                    press_No,
                    Wrok_time,
                    selected_Work_Item,
                    selected_Work,
                    comment,
                    key_number
                )
            db_sys.close()
            cols[1].info("登録しました。")
        elif button1 == True and entry == True and CB == False :
            cols[1].error("※ 確認のためチェックをしてください")
        elif button1 == True and entry == False :
            cols[1].error("※ 全て記入してからの登録をお願いします。")
        # ----------------------------------------------------
#=================================================================================================================

    elif st.session_state['pages'] == "経過時間表図化":
        # -------------------------------------------------------------------------------------
        # Sidebar
        # -------------------------------------------------------------------------------------
        password_No = st.sidebar.text_input("■ パスワードを記入してください。",type='password')
        login_button = st.sidebar.button("ログイン")
        logout_button = st.sidebar.button("ログアウト")

        #パスワード設定（谷本さんのみ）-----
        password_setting = "tanimoto"
        #---------------------------------

        if 'password' not in st.session_state :
            st.session_state['password'] = ""            #保持変数（password）の初期値
        if 'limit' not in st.session_state :
            st.session_state['limit'] = 3                #保持変数（limit）の初期値
        elif password_No == password_setting and login_button == True:
            if st.session_state['limit'] > 0 :           #パスワード入力回数カウント
                st.session_state['limit'] = 3            #パスワード入力回数のリセット
                st.session_state['password'] = "成功"  #保持変数に（成功！）を代入
        elif password_No != password_setting and login_button == True:
            if st.session_state['limit'] > 0 :            #パスワード入力回数カウント
                st.session_state['limit'] -= 1            #パスワード入力回数の減少
                st.session_state['password'] = "失敗" #保持変数に（失敗；；）を代入
                if st.session_state['limit'] == 0 :
                    st.session_state['password'] = "リセット"

        if st.session_state['password'] == "成功" and logout_button == True:
            st.session_state['password'] = ""
            st.session_state['limit'] = 3

        st.sidebar.write(f"パスワード入力チャンスはあと{st.session_state['limit']}回です。")

        if st.session_state['password'] == "":
            st.sidebar.write("<b>not login(·ε·｀)ｽﾈﾁｬｳ</b>",unsafe_allow_html=True)
        elif st.session_state['password'] == "成功":
            st.sidebar.success("ログイン成功ヽ(*´∀｀)ノ")
        elif st.session_state['password'] == "失敗":
            st.sidebar.error("ログイン失敗( ´•̥̥̥ω•̥̥̥` )")
        elif st.session_state['password'] == "リセット":
            st.sidebar.warning("開発者を呼んでください。( ﾟДﾟ)㌦ｧ!!")
        
        # -------------------------------------------------------------------------------------
        # -------------------------------------------------------------------------------------
        if st.session_state['password'] == "成功":
            st.write("<b>グラフ</b>",unsafe_allow_html=True)






class Work_sheet_Database:
    # 初期化=====================================================================
    def __init__(self):
        self.conn   = sqlite3.connect(BUP_DB_Path, check_same_thread=False)
        self.cur    = self.conn.cursor()
        #self.table  = []
    #===========================================================================
    
    # データ書き込み=============================================================
    def get(self, workers_name, date, selected_Press, selected_Press_type,\
            press_No, Wrok_time, selected_Work_Item, selected_Work, comment,\
            key_number):

        db_path = '//192.168.1.212/アイシス/00_製造_自動発注システム/20_DataBase/main.db'
        # DBを作成する（既に作成されていたらこのDBに接続する）
        conn = sqlite3.connect(db_path)
        # SQLiteを操作するためのカーソルを作成
        cur = conn.cursor()

        #PLENOX機種内の「-」がSQL INSERT時にエラーが出てしまう為。
        db_selected_Press_type = selected_Press_type.replace('-', 'ー')       
        db_date = date.strftime('%Y/%m/%d')

        # テーブルの作成
        table_name = str(selected_Press) + "_" + str(db_selected_Press_type) + "_" + str(press_No)
        sql_create = "CREATE TABLE IF NOT EXISTS "
        sql_columns = """(id INTEGER PRIMARY KEY AUTOINCREMENT,
                          workers_name TEXT,   date TEXT,
                          Wrok_time TEXT,      selected_Work_Item TEXT,
                          selected_Work TEXT,  comment TEXT
                          )"""
        sql_C_table = sql_create + table_name + sql_columns
        cur.execute(sql_C_table)

        # SQLに保存
        sql_insert_front = "INSERT INTO "
        sql_insert_rear = """(workers_name,   date,
                              Wrok_time,      selected_Work_Item,
                              selected_Work,  comment
                              )
                              values (?,?,?,?,?,?)"""
        sql_insert = sql_insert_front + table_name + sql_insert_rear
        for R_FA in range(key_number-1):
            insert_date = workers_name,          db_date,\
                          str(Wrok_time[R_FA]),  selected_Work_Item[R_FA],  \
                          selected_Work[R_FA],   comment[R_FA]
            cur.execute(sql_insert, insert_date)
            conn.commit()

    #===========================================================================
    
    # SQLを閉じる================================================================
    def close(self):
        self.cur.close()
        self.conn.close()
    #===========================================================================
        
start_Order_His()